from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List
from datetime import datetime
import os

def generar_excel_reportes(reportes: List, nombre_archivo: str = None) -> str:

    if nombre_archivo is None:
        fecha = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre_archivo = f"reportes_incidencias_{fecha}.xlsx"
    else:
        nombre_archivo = f"{nombre_archivo}.xlsx"
    # Crear carpeta si no existe
    ruta_carpeta = os.path.join("uploads", "reportes_excel")
    os.makedirs(ruta_carpeta, exist_ok=True)
    ruta_completa = os.path.join(ruta_carpeta, nombre_archivo)
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Reportes"
    
    # Definir estilos
    encabezado_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    encabezado_font = Font(bold=True, color="FFFFFF", size=12)
    encabezado_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    borde = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Definir columnas
    columnas = [
        "Folio",
        "Fecha",
        "Tipo de Problema",
        "Edificio",
        "Nivel",
        "Sexo",
        "Pasillo",
        "Taza/Orinal",
        "Número de Cuenta",
        "Prioridad",
        "Estado",
        "Observaciones"
    ]
    
    # Escribir encabezados
    for col_num, titulo_columna in enumerate(columnas, 1):
        celda = ws.cell(row=1, column=col_num)
        celda.value = titulo_columna
        celda.fill = encabezado_fill
        celda.font = encabezado_font
        celda.alignment = encabezado_alignment
        celda.border = borde
    
    # Ajustar altura de encabezado
    ws.row_dimensions[1].height = 25
    
    # Llenar datos
    for row_num, reporte in enumerate(reportes, 2):
        datos = [
            getattr(reporte, 'folio', ''),
            getattr(reporte, 'fecha_creacion', ''),
            getattr(reporte, 'tipo_problema', ''),
            getattr(reporte, 'edificio', ''),
            getattr(reporte, 'nivel', ''),
            getattr(reporte, 'sexo', ''),
            getattr(reporte, 'pasillo', ''),
            getattr(reporte, 'taza_o_orinal', ''),
            getattr(reporte, 'numero_cuenta', ''),
            getattr(reporte, 'prioridad_asignada', ''),
            getattr(reporte, 'estado', ''),
            getattr(reporte, 'observaciones', '')
        ]
        
        for col_num, valor in enumerate(datos, 1):
            celda = ws.cell(row=row_num, column=col_num)
            celda.value = valor
            celda.alignment = Alignment(horizontal="left", vertical="center")
            celda.border = borde
    
    # Ajustar ancho de columnas
    anchos = [15, 15, 20, 15, 8, 8, 12, 12, 15, 12, 12, 20]
    for col_num, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = ancho
    
    # Agregar resumen al final
    fila_resumen = len(reportes) + 3
    ws.cell(row=fila_resumen, column=1).value = f"Total de reportes: {len(reportes)}"
    ws.cell(row=fila_resumen, column=1).font = Font(bold=True)
    
    ws.cell(row=fila_resumen + 1, column=1).value = f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    
    # Guardar archivo
    wb.save(ruta_completa)
    
    return ruta_completa
